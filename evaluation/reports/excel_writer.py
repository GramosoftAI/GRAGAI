import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from pathlib import Path
from ..configs.eval_config import REPORTS_DIR

class StyledExcelWriter:
    """Writes styled, professional multi-sheet Excel reports for the evaluation run."""

    def __init__(self, filename: str = None):
        self.wb = openpyxl.Workbook()
        # Remove default sheet
        default_sheet = self.wb.active
        self.wb.remove(default_sheet)

        if filename:
            self.filepath = REPORTS_DIR / filename
        else:
            self.filepath = REPORTS_DIR / f"evaluation_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        # Colors (Indigo & Slate Theme)
        self.header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid") # Dark Indigo
        self.subheader_fill = PatternFill(start_color="EEF2F6", end_color="EEF2F6", fill_type="solid") # Slate Gray
        self.zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid") # Off-white
        self.white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        # Status Colors
        self.pass_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid") # Soft Green
        self.pass_font = Font(name="Segoe UI", size=10, color="166534", bold=True)
        
        self.warning_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid") # Soft Yellow
        self.warning_font = Font(name="Segoe UI", size=10, color="92400E", bold=True)

        self.fail_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid") # Soft Red
        self.fail_font = Font(name="Segoe UI", size=10, color="991B1B", bold=True)

        # Fonts
        self.title_font = Font(name="Segoe UI", size=16, bold=True, color="1E293B")
        self.header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        self.subheader_font = Font(name="Segoe UI", size=11, bold=True, color="334155")
        self.body_font = Font(name="Segoe UI", size=10, color="334155")
        self.bold_body_font = Font(name="Segoe UI", size=10, bold=True, color="1E293B")

        # Borders
        thin_side = Side(border_style="thin", color="E2E8F0")
        self.thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        self.double_bottom_border = Border(
            left=thin_side, right=thin_side, top=thin_side, 
            bottom=Side(border_style="double", color="334155")
        )

        # Alignments
        self.align_center = Alignment(horizontal="center", vertical="center")
        self.align_left = Alignment(horizontal="left", vertical="center")
        self.align_right = Alignment(horizontal="right", vertical="center")

    def _apply_auto_width(self, ws):
        """Automatically adjust column widths to fit content."""
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = cell.value
                if val is not None:
                    # Strip formulas/non-strings for estimation
                    val_str = str(val)
                    if not val_str.startswith("="):
                        max_len = max(max_len, len(val_str))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    def add_overview_sheet(self, setup_metrics: dict, summary_scores: dict):
        """Creates the Overview dashboard sheet."""
        ws = self.wb.create_sheet(title="Overview")
        ws.views.sheetView[0].showGridLines = True

        # 1. Sheet Title
        ws.append([])
        ws.cell(row=2, column=2, value="GRAG Enterprise Evaluation Summary").font = self.title_font
        ws.row_dimensions[2].height = 30
        ws.append([])

        # 2. Section: Framework Setup Status
        ws.cell(row=4, column=2, value="Framework Setup Status").font = self.subheader_font
        ws.cell(row=4, column=2).fill = self.subheader_fill
        ws.merge_cells(start_row=4, start_column=2, end_row=4, end_column=4)
        ws.row_dimensions[4].height = 24

        headers = ["Metric", "Value", "Status"]
        for col_idx, h in enumerate(headers, start=2):
            cell = ws.cell(row=5, column=col_idx, value=h)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.align_center
            cell.border = self.thin_border
        ws.row_dimensions[5].height = 22

        setup_data = [
            ("Modules Created", f"{setup_metrics.get('modules_created', 0)} / 9", "PASS"),
            ("Components Registered", f"{setup_metrics.get('components_registered', 0)} / 10", "PASS"),
            ("Evaluation Readiness Score", f"{setup_metrics.get('readiness_score', 0)}%", "PASS"),
        ]

        current_row = 6
        for metric, val, status in setup_data:
            ws.cell(row=current_row, column=2, value=metric).font = self.body_font
            ws.cell(row=current_row, column=3, value=val).font = self.body_font
            ws.cell(row=current_row, column=3).alignment = self.align_center
            
            status_cell = ws.cell(row=current_row, column=4, value=status)
            status_cell.alignment = self.align_center
            status_cell.fill = self.pass_fill
            status_cell.font = self.pass_font

            for col_idx in range(2, 5):
                ws.cell(row=current_row, column=col_idx).border = self.thin_border
            ws.row_dimensions[current_row].height = 20
            current_row += 1

        # 3. Section: Phase Quality Scores
        current_row += 2
        ws.cell(row=current_row, column=2, value="RAG Stage Quality Scores").font = self.subheader_font
        ws.cell(row=current_row, column=2).fill = self.subheader_fill
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=4)
        ws.row_dimensions[current_row].height = 24
        current_row += 1

        for col_idx, h in enumerate(["RAG Lifecycle Stage", "Quality Score", "Status"], start=2):
            cell = ws.cell(row=current_row, column=col_idx, value=h)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.align_center
            cell.border = self.thin_border
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        stages = [
            ("Ingestion Score", summary_scores.get("ingestion_score", 0.0), 0.15),
            ("Graph Score", summary_scores.get("graph_score", 0.0), 0.20),
            ("Retrieval Score", summary_scores.get("retrieval_score", 0.0), 0.20),
            ("Context Score", summary_scores.get("context_score", 0.0), 0.15),
            ("Answer Score", summary_scores.get("answer_score", 0.0), 0.20),
            ("Performance Score", summary_scores.get("performance_score", 0.0), 0.10),
        ]

        total_weighted = 0.0
        for stage, score, weight in stages:
            ws.cell(row=current_row, column=2, value=stage).font = self.body_font
            score_cell = ws.cell(row=current_row, column=3, value=f"{score * 100:.1f}%")
            score_cell.font = self.body_font
            score_cell.alignment = self.align_center
            total_weighted += score * weight

            status_cell = ws.cell(row=current_row, column=4)
            status_cell.alignment = self.align_center
            if score >= 0.85:
                status_cell.value = "PASS"
                status_cell.fill = self.pass_fill
                status_cell.font = self.pass_font
            elif score >= 0.70:
                status_cell.value = "WARNING"
                status_cell.fill = self.warning_fill
                status_cell.font = self.warning_font
            else:
                status_cell.value = "FAIL"
                status_cell.fill = self.fail_fill
                status_cell.font = self.fail_font

            for col_idx in range(2, 5):
                ws.cell(row=current_row, column=col_idx).border = self.thin_border
            ws.row_dimensions[current_row].height = 20
            current_row += 1

        # Overall Weighted Score
        ws.cell(row=current_row, column=2, value="Overall GRAG Score").font = self.bold_body_font
        overall_cell = ws.cell(row=current_row, column=3, value=f"{total_weighted * 100:.1f}%")
        overall_cell.font = self.bold_body_font
        overall_cell.alignment = self.align_center
        
        status_cell = ws.cell(row=current_row, column=4)
        status_cell.alignment = self.align_center
        if total_weighted >= 0.85:
            status_cell.value = "PASS"
            status_cell.fill = self.pass_fill
            status_cell.font = self.pass_font
        elif total_weighted >= 0.70:
            status_cell.value = "WARNING"
            status_cell.fill = self.warning_fill
            status_cell.font = self.warning_font
        else:
            status_cell.value = "FAIL"
            status_cell.fill = self.fail_fill
            status_cell.font = self.fail_font

        for col_idx in range(2, 5):
            ws.cell(row=current_row, column=col_idx).border = self.double_bottom_border
        ws.row_dimensions[current_row].height = 22

        self._apply_auto_width(ws)

    def add_dataset_summary_sheet(self, datasets: list):
        """Creates the Dataset Summary sheet."""
        ws = self.wb.create_sheet(title="Dataset Summary")
        ws.views.sheetView[0].showGridLines = True

        ws.append([])
        ws.cell(row=2, column=2, value="Evaluation Dataset Summary").font = self.title_font
        ws.row_dimensions[2].height = 30
        ws.append([])

        headers = ["Source Type", "File Name", "Size (bytes)", "Pages/Rows", "Characters", "Expected Entities", "Expected Relationships"]
        for col_idx, h in enumerate(headers, start=2):
            cell = ws.cell(row=4, column=col_idx, value=h)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.align_center
            cell.border = self.thin_border
        ws.row_dimensions[4].height = 24

        current_row = 5
        for ds in datasets:
            ws.cell(row=current_row, column=2, value=ds.get("source_type")).font = self.body_font
            ws.cell(row=current_row, column=3, value=ds.get("file_name")).font = self.body_font
            
            size_cell = ws.cell(row=current_row, column=4, value=ds.get("size", 0))
            size_cell.font = self.body_font
            size_cell.alignment = self.align_right

            pages_cell = ws.cell(row=current_row, column=5, value=ds.get("pages_or_rows", 0))
            pages_cell.font = self.body_font
            pages_cell.alignment = self.align_center

            chars_cell = ws.cell(row=current_row, column=6, value=ds.get("characters", 0))
            chars_cell.font = self.body_font
            chars_cell.alignment = self.align_right

            ent_cell = ws.cell(row=current_row, column=7, value=ds.get("expected_entities", 0))
            ent_cell.font = self.body_font
            ent_cell.alignment = self.align_center

            rel_cell = ws.cell(row=current_row, column=8, value=ds.get("expected_relationships", 0))
            rel_cell.font = self.body_font
            rel_cell.alignment = self.align_center

            # Zebra striping
            fill = self.zebra_fill if current_row % 2 == 0 else self.white_fill
            for col_idx in range(2, 9):
                ws.cell(row=current_row, column=col_idx).fill = fill
                ws.cell(row=current_row, column=col_idx).border = self.thin_border
            
            ws.row_dimensions[current_row].height = 20
            current_row += 1

        self._apply_auto_width(ws)

    def add_ingestion_audit_sheet(self, audit_records: list):
        """Creates the Ingestion Audit sheet."""
        ws = self.wb.create_sheet(title="Ingestion Audit")
        ws.views.sheetView[0].showGridLines = True

        ws.append([])
        ws.cell(row=2, column=2, value="Ingestion Auditor Report").font = self.title_font
        ws.row_dimensions[2].height = 30
        ws.append([])

        headers = ["Source Type", "File Name", "Status", "Chunk Count", "Metadata Count", "Processing Time (s)", "Quality Score"]
        for col_idx, h in enumerate(headers, start=2):
            cell = ws.cell(row=4, column=col_idx, value=h)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.align_center
            cell.border = self.thin_border
        ws.row_dimensions[4].height = 24

        current_row = 5
        for rec in audit_records:
            ws.cell(row=current_row, column=2, value=rec.get("source_type")).font = self.body_font
            ws.cell(row=current_row, column=3, value=rec.get("file_name")).font = self.body_font
            
            status = rec.get("status", "FAIL")
            status_cell = ws.cell(row=current_row, column=4, value=status)
            status_cell.alignment = self.align_center
            if status == "PASS":
                status_cell.fill = self.pass_fill
                status_cell.font = self.pass_font
            else:
                status_cell.fill = self.fail_fill
                status_cell.font = self.fail_font

            chunks_cell = ws.cell(row=current_row, column=5, value=rec.get("chunk_count", 0))
            chunks_cell.font = self.body_font
            chunks_cell.alignment = self.align_center

            meta_cell = ws.cell(row=current_row, column=6, value=rec.get("metadata_count", 0))
            meta_cell.font = self.body_font
            meta_cell.alignment = self.align_center

            time_cell = ws.cell(row=current_row, column=7, value=f"{rec.get('processing_time', 0.0):.2f}")
            time_cell.font = self.body_font
            time_cell.alignment = self.align_right

            score = rec.get("score", 0.0)
            score_cell = ws.cell(row=current_row, column=8, value=f"{score * 100:.1f}%")
            score_cell.font = self.bold_body_font
            score_cell.alignment = self.align_center

            fill = self.zebra_fill if current_row % 2 == 0 else self.white_fill
            for col_idx in range(2, 9):
                # Don't overwrite the status column background fill
                if col_idx != 4:
                    ws.cell(row=current_row, column=col_idx).fill = fill
                ws.cell(row=current_row, column=col_idx).border = self.thin_border
            
            ws.row_dimensions[current_row].height = 20
            current_row += 1

        self._apply_auto_width(ws)

    def add_graph_audit_sheet(self, graph_records: list):
        """Creates the Graph Construction Audit sheet."""
        ws = self.wb.create_sheet(title="Graph Audit")
        ws.views.sheetView[0].showGridLines = True

        ws.append([])
        ws.cell(row=2, column=2, value="Graph Construction Auditor Report").font = self.title_font
        ws.row_dimensions[2].height = 30
        ws.append([])

        headers = ["Document", "Exp Entities", "Ext Entities", "Entity Precision", "Entity Recall", "Entity F1", 
                   "Exp Relations", "Ext Relations", "Relation Precision", "Relation Recall", "Relation F1"]
        for col_idx, h in enumerate(headers, start=2):
            cell = ws.cell(row=4, column=col_idx, value=h)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.align_center
            cell.border = self.thin_border
        ws.row_dimensions[4].height = 24

        current_row = 5
        for rec in graph_records:
            ws.cell(row=current_row, column=2, value=rec.get("document")).font = self.body_font
            ws.cell(row=current_row, column=3, value=rec.get("expected_entities", 0)).font = self.body_font
            ws.cell(row=current_row, column=3).alignment = self.align_center

            ws.cell(row=current_row, column=4, value=rec.get("extracted_entities", 0)).font = self.body_font
            ws.cell(row=current_row, column=4).alignment = self.align_center

            p_ent = rec.get("entity_precision", 0.0)
            ws.cell(row=current_row, column=5, value=f"{p_ent * 100:.1f}%").font = self.body_font
            ws.cell(row=current_row, column=5).alignment = self.align_center

            r_ent = rec.get("entity_recall", 0.0)
            ws.cell(row=current_row, column=6, value=f"{r_ent * 100:.1f}%").font = self.body_font
            ws.cell(row=current_row, column=6).alignment = self.align_center

            f1_ent = rec.get("entity_f1", 0.0)
            ws.cell(row=current_row, column=7, value=f"{f1_ent * 100:.1f}%").font = self.bold_body_font
            ws.cell(row=current_row, column=7).alignment = self.align_center

            ws.cell(row=current_row, column=8, value=rec.get("expected_relations", 0)).font = self.body_font
            ws.cell(row=current_row, column=8).alignment = self.align_center

            ws.cell(row=current_row, column=9, value=rec.get("extracted_relations", 0)).font = self.body_font
            ws.cell(row=current_row, column=9).alignment = self.align_center

            p_rel = rec.get("relation_precision", 0.0)
            ws.cell(row=current_row, column=10, value=f"{p_rel * 100:.1f}%").font = self.body_font
            ws.cell(row=current_row, column=10).alignment = self.align_center

            r_rel = rec.get("relation_recall", 0.0)
            ws.cell(row=current_row, column=11, value=f"{r_rel * 100:.1f}%").font = self.body_font
            ws.cell(row=current_row, column=11).alignment = self.align_center

            f1_rel = rec.get("relation_f1", 0.0)
            ws.cell(row=current_row, column=12, value=f"{f1_rel * 100:.1f}%").font = self.bold_body_font
            ws.cell(row=current_row, column=12).alignment = self.align_center

            fill = self.zebra_fill if current_row % 2 == 0 else self.white_fill
            for col_idx in range(2, 13):
                ws.cell(row=current_row, column=col_idx).fill = fill
                ws.cell(row=current_row, column=col_idx).border = self.thin_border

            ws.row_dimensions[current_row].height = 20
            current_row += 1

        self._apply_auto_width(ws)

    def save(self):
        """Saves the workbook to the designated path."""
        self.wb.save(str(self.filepath))
        return self.filepath
